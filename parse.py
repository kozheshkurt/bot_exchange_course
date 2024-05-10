import time
from config import EXCEL_FILENAME, DATABASE_FILENAME

from selenium import webdriver
from selenium.webdriver.common.by import By

import sqlite3

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill


# функція парсить запропонований сайт та повертає поточний курс
def get_exchange_rate():
    driver = webdriver.Chrome()
    driver.get("https://www.google.com/finance/quote/USD-UAH")

    element = driver.find_element(By.CLASS_NAME,'fxKbKc') # знаходимо по назві класу, яку попередньо знайшов на сторінці
    course = element.text # зберіг в окремій змінній, щоб повернути значення після закриття драйверу
    
    driver.close()

    # print(f'Поточний курс USD/UAH: {course}')
    return course


# функція в поточний момент часу (який є аргументом) формує кортеж із пари1 значень для внесення в БД. Оскільки нам потрібні значення часу HH:00:00, це момент спрацювання таймеру в циклі while, і цей момент задається як аргумент 
def get_data_for_sql(current_time):
    current_time_sql = time.strftime('%Y-%m-%d %H:%M:%S', current_time) # перетворюємо час у формат, в якому зберігаємо час в SQLite як str
    rate = get_exchange_rate()
    return current_time_sql, rate


# функція створює файл бази даних, якщо це перший запуск програми і файлу ще немає
def database_start():
    db = sqlite3.connect(DATABASE_FILENAME)
    cursor = db.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS exchange_rates (
            id INTEGER PRIMARY KEY,
            datetime TEXT NOT NULL,
            rate REAL NOT NULL
        )
    ''')
    db.commit()
    db.close()

# функція приймає значення часу (у форматі для зберігання в БД) та курсу та додає запис в базу даних. Пару значень формує функція get_data_for_sql
def database_add_rate(datetime_sql, rate):
    db = sqlite3.connect(DATABASE_FILENAME)
    cursor = db.cursor()

    cursor.execute('INSERT INTO exchange_rates (datetime, rate) VALUES (?, ?)', (datetime_sql, rate))

    db.commit()
    db.close()


# отримуємо значення курсів за сьогоднішню (фільтр WHILE) дату. Повертає список кортежів (датачас, курс)
def database_get_data_for_excel():
    db = sqlite3.connect(DATABASE_FILENAME)
    cursor = db.cursor()
    
    cursor.execute('''
        SELECT strftime('%d.%m.%Y %H:%M:%S', datetime), rate
        FROM exchange_rates
        WHERE date('now') = date(datetime)
    ''')

    today_rates = cursor.fetchall()

    db.commit()
    db.close()

    return today_rates


# функція створює файл excel з потрібним форматуванням при першому запуску програми
def excel_start():
    try:
        wb = openpyxl.load_workbook(EXCEL_FILENAME) # якщо файл відкривається, нічого не робимо
        wb.close()
    except:
        wb = Workbook()
        sheet = wb.active  # якщо файл не відкривається, створюємо новий
        
        sheet.column_dimensions['A'].width = 20     # задаємо ширину перших двох колонок
        sheet.column_dimensions['B'].width = 15
        
        fill = PatternFill(fill_type='solid', fgColor='00FFFF00') # зафарбовуємо "шапку" жовтим
        sheet['A1'].fill = fill
        sheet['B1'].fill = fill
        
        wb.save(EXCEL_FILENAME)
        wb.close()

        # заповнюємо створений файл согоднішніми курсами валют
        today_rates = database_get_data_for_excel()
        update_rates_in_excel(EXCEL_FILENAME, today_rates) 



# функція перезаписує xlsx файл, отримуючи актуальні курси, отримані із SQLite у вигляді списку кортежів (датачас, курс)
def update_rates_in_excel(filename, today_rates):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    # цикл видаляє всі значення з таблиці. Не додає до існуючих, тому що там можуть бути вчорашні курси, і ми заповнимо таблицю з нуля сьогоднішніми
    for row in sheet.iter_rows(min_row=1, max_col = 2, max_row = sheet.max_row): # колонок лише дві, рядків - скільки є заповнених
        for cell in row:
            cell.value = None

    wb.save(filename)
    wb.close() # файл закриваємо та відкриваємо знову, тому що без цього sheet.append(rate) прикріплює нові значення нижче останнього, який був заповнений до того

    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    sheet['A1'], sheet['B1'] = 'datetime', 'exchange_rate' # шапка таблиці
    for rate in today_rates:
        sheet.append(rate)

    wb.save(filename)
    wb.close()


# початок роботи, створення файлів, якщо їх ще немає
database_start()
excel_start()

while True:
    time.sleep(1)
    current_time = time.localtime()    # кожну секунду перевіряємо, чи є кількість годин круглою (НН:00:00). Цей час зафіксували у змінній current_time і використовуємо його надалі
   
    # if current_time.tm_sec % 30 == 0: # для тестування, виконує оновлення кожні 30 секунд
    if current_time.tm_min == 0 and current_time.tm_sec == 0: 
        
        current_time_sql, rate = get_data_for_sql(current_time) # отримуємо значення часу та курсу
        database_add_rate(current_time_sql, rate) # додаємо значення часу та курсу в БД

        today_rates = database_get_data_for_excel() # отримуємо значення часу та курсу на сьогоднішню дату
        update_rates_in_excel(EXCEL_FILENAME, today_rates) # оновлюємо таблицю xlsx, готову до відправки користувачу
